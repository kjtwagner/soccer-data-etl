import random
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

df = pd.read_csv('names.csv', header=None)
players = df[0].tolist()
print(players)

goalkeepers = ["Maddie", "Emily", "Karissa", "Cam C", "Ada"]
random.shuffle(goalkeepers)

num_rounds = 5
num_teams = 4
colors = ['Pink', 'Blue', 'Yellow', 'Jerseys']

round_dfs = []
player_records = {name: [] for name in players + goalkeepers}

for round_num in range(1, num_rounds + 1):
    # shuffle em up
    shuffled = players[:] # since random mutates orig list
    random.shuffle(shuffled)

    # handle GKs
    excluded_keeper = goalkeepers[round_num - 1]
    active_keepers = [g for g in goalkeepers if g != excluded_keeper]
    random.shuffle(active_keepers)

    # build teams
    teams = [[] for _ in range(num_teams)]
    
    for i in range(num_teams):
        teams[i].append(active_keepers[i])
        
    for i, player in enumerate(shuffled):
        teams[i % num_teams].append(player)
    
    # record team by player
    for team_idx, team in enumerate(teams):
        for p in team:
            player_records[p].append(colors[team_idx])
    player_records[excluded_keeper].append("Rest")

    # zero pad for alignment
    max_len = max(len(team) for team in teams)
    for i in range(num_teams):
        teams[i] += [""] * (max_len - len(teams[i]))

    round_data = {f"{colors[i]}": teams[i] for i in range(num_teams)}
    round_dfs.append(pd.DataFrame(round_data))

## ROUND centered output ##
# try excel formatting
out_file = "scrimmage_teams.xlsx"
with pd.ExcelWriter(out_file, engine="openpyxl") as writer:
    startrow = 0
    for rd in round_dfs:
        rd.to_excel(writer, sheet_name="Teams", startrow=startrow, index=False)
        startrow += len(rd) + 2

wb = load_workbook(out_file)
ws = wb["Teams"]

header_fill = PatternFill(start_color="a3a29f", end_color="a3a29f", fill_type="solid")
header_font = Font(bold=True)

# highlight header rows
for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
    if any(cell.value and "Pink" in str(cell.value) for cell in row):
        for cell in row:
            cell.fill = header_fill
            cell.font = header_font
            
wb.save(out_file)

## PLAYER/GK centered output ##
columns = []
for r in range(1, num_rounds+1):
    columns += ["Team", "Team Pts", "Goals", f"Rd{r} Total"]

data_players = []
for player in players:
    row = [player]
    for r in range(num_rounds):
        team = player_records[player][r] if r < len(player_records[player]) else ""
        row += [team, "", "", ""]
    data_players.append(row)

df_players = pd.DataFrame(data_players, columns=["Name"] + columns)
df_players.to_excel("round_summary_players.xlsx", index=False)

data_gks = []
for gk in goalkeepers:
    row = [gk]
    for r in range(num_rounds):
        team = player_records[gk][r] if r < len(player_records[gk]) else ""
        row += [team, "", "", ""]
    data_gks.append(row)

df_gks = pd.DataFrame(data_gks, columns=["Name"] + columns)
df_gks.to_excel("round_summary_gks.xlsx", index=False)
